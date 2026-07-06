import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SC = sys.argv[1]; FOLDER = sys.argv[2]
rk = json.load(open(f"{SC}/theme_ranking.json"))

def docpath(name): return f"{FOLDER}\\{name}.docx"
DOCS = {
 '00':'DOC 00 — Master Spine (Veredicto Estratégico)',
 '15':'DOC 15 — Retrato del Cliente (Avatar Core 5 + Sub-avatares)',
 '16':'DOC 16 — Banco de Lenguaje, Hooks e Identidad Reclamada',
 '19':'DOC 19 — UMP-UMS-USP Mecanismos Únicos',
 '20':'DOC 20 — Mapa de Dolores + Buyer Psychology',
 'F0':'FASE 0 — Higado+Intestino MX (Purelia)',
}
LABELS = {
'bloating':'Hinchazón / panza inflamada','fatty_liver':'Hígado graso / NAFLD','belief_natural':'Preferencia por lo natural',
'toxin_detox':'Desintoxicar / detox','liver_enzymes':'Enzimas hepáticas (ALT/AST)','heavy_digestion':'Digestión pesada / lenta',
'weight_stubborn':'Peso que no baja','desire_flat_stomach':'Deseo: vientre plano','bile_flow':'Flujo biliar / digestión de grasas',
'alcohol_liver':'Alcohol → hígado','obj_does_it_work':'Objeción: ¿funciona?','constipation':'Estreñimiento',
'belief_its_age':'Creencia: es la edad','obj_taste':'Objeción: sabor de gotas','chronic_fatigue':'Cansancio crónico',
'gas_pain':'Gases / dolor abdominal','brain_fog':'Niebla mental','energy_crash_2pm':'Bajón de las 2pm','skin_dull':'Piel opaca / granitos',
'desire_energy':'Deseo: energía','desire_feel_myself':'Deseo: volver a sentirme yo','obj_scam':'Objeción: ¿otra estafa?',
'obj_price':'Objeción: precio','obj_tried_everything':'Objeción: ya probé todo','belief_self_blame':'Autoculpa (dieta)',
'belief_doctor_dismissed':'El doctor no me ayudó','identity_lost_self':'Identidad: la que era antes','identity_ashamed':'Identidad: vergüenza del cuerpo',
}
def tier(pi):
    return 'S' if pi>=90 else 'A' if pi>=80 else 'B' if pi>=65 else 'C'
TIERFILL = {'S':'F4B942','A':'A9D08E','B':'FFE699','C':'F4B183'}  # gold/green/yellow/orange

NAVY='1F3864'; HDR=Font(bold=True,color='FFFFFF',name='Calibri',size=11)
HFILL=PatternFill('solid',fgColor=NAVY)
TITLE=Font(bold=True,size=16,color=NAVY,name='Calibri')
SUB=Font(bold=True,size=12,color=NAVY,name='Calibri')
thin=Side(style='thin',color='BFBFBF'); BORD=Border(thin,thin,thin,thin)
WRAP=Alignment(wrap_text=True,vertical='top')
CTR=Alignment(horizontal='center',vertical='center')

wb=Workbook()

def style_header(ws,row,ncols):
    for c in range(1,ncols+1):
        cell=ws.cell(row=row,column=c); cell.font=HDR; cell.fill=HFILL; cell.alignment=Alignment(wrap_text=True,vertical='center'); cell.border=BORD
def borders(ws,r1,r2,ncols):
    for r in range(r1,r2+1):
        for c in range(1,ncols+1):
            cell=ws.cell(row=r,column=c); cell.border=BORD
            if not cell.alignment.wrap_text: cell.alignment=WRAP
def tiercell(ws,row,col,t):
    cell=ws.cell(row=row,column=col,value=t); cell.fill=PatternFill('solid',fgColor=TIERFILL[t]); cell.alignment=CTR; cell.font=Font(bold=True); cell.border=BORD
def link(ws,row,col,docid,text='📄 Ver documento completo'):
    cell=ws.cell(row=row,column=col,value=text); cell.hyperlink=docpath(DOCS[docid]); cell.font=Font(color='0563C1',underline='single'); cell.border=BORD

# ---------- TAB 1: RESUMEN ----------
ws=wb.active; ws.title='Resumen'
ws['A1']='RESEARCH CONSOLIDADO — Suplemento Hepático · Lanzamiento MÉXICO'; ws['A1'].font=TITLE
ws['A2']='Corpus real: 13,548 confesiones · 6 plataformas × 2 idiomas · Motor de Research v3'; ws['A2'].font=Font(italic=True,color='808080')
r=4
ws.cell(r,1,'PRODUCTO').font=SUB; r+=1
prod=[('Producto','Suplemento de salud hepática en gotas líquidas (base competidor: Purelia Complete Liver Support)'),
('Ingredientes','L-Glutatión, NAC, Cardo Mariano (silimarina), Diente de León, Alcachofa, Betabel, Colina'),
('Formato','Gotas líquidas sublinguales — ángulo "hasta 98% absorción" vs cápsulas'),
('Mercado','Lanzamiento México (es) · Inteligencia EE.UU. (en) · Categoría: hígado + intestino')]
for k,v in prod:
    ws.cell(r,1,k).font=Font(bold=True); ws.cell(r,2,v); r+=1
r+=1
ws.cell(r,1,'VEREDICTO EN UNA LÍNEA').font=SUB; r+=1
ws.cell(r,1,'A la mujer mexicana de 40+ que "hace todo bien" y aún así vive hinchada, le vendemos volver a sentirse ligera y ella misma — destapándole el hígado (no poniéndola a dieta) — con la hinchazón como gancho y las enzimas como prueba.')
ws.cell(r,1).alignment=WRAP; ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); ws.row_dimensions[r].height=45; r+=2
ws.cell(r,1,'RECOMENDACIÓN GO-TO-MARKET (cómo salir al mercado)').font=SUB; r+=1
gtm=[('1. Apuntar a','Mujer 40+ hinchada que "hace todo bien"'),
('2. Gancho','La hinchazón / panza (NO las enzimas hepáticas)'),
('3. Gran idea','"No es el pan ni tu disciplina — es tu HÍGADO TAPADO"'),
('4. Producto','El "destape en gotas" naturales'),
('5. Promesa','Panza plana y ligereza SIN dieta'),
('6. Villano','Frituras / azúcar / comida pesada (NO alcohol)'),
('7. Prueba','Enzimas ALT/AST + testimonios'),
('8. Tono','Alivio + des-culpabilización, natural, cero "cura milagro" (compliance)')]
for k,v in gtm:
    ws.cell(r,1,k).font=Font(bold=True); ws.cell(r,2,v); r+=1
r+=1
ws.cell(r,1,'ÍNDICE DE DOCUMENTOS (clic para abrir)').font=SUB; r+=1
idx=[('00','Master Spine — el veredicto consolidado'),('15','Avatar Core 5 + Sub-avatares'),
('16','Banco de Lenguaje, Hooks e Identidad Reclamada'),('19','UMP / UMS / USP — Mecanismos Únicos'),
('20','Mapa de Dolores + Buyer Psychology'),('F0','Fase 0 — Competencia y vacíos de mercado')]
for did,desc in idx:
    link(ws,r,1,did,f'📄 DOC {did if did!="F0" else "FASE 0"} — {desc}'); r+=1
ws.column_dimensions['A'].width=34; ws.column_dimensions['B'].width=70
for c in 'CDEF': ws.column_dimensions[c].width=14

# ---------- TAB 2: COMPETENCIA ----------
ws=wb.create_sheet('Competencia')
ws['A1']='COMPETENCIA PRINCIPAL'; ws['A1'].font=TITLE
link(ws,2,1,'F0','📄 Fuente: FASE 0 — Competencia y vacíos')
hdr=['Competidor','Segmento','Ángulo / Gran idea','Mercado','Debilidad','Relevancia']
hr=4
for c,h in enumerate(hdr,1): ws.cell(hr,c,h)
style_header(ws,hr,len(hdr))
comp=[
('Purelia (base)','Hígado','"Tu hígado es la raíz de tu bajón, hinchazón y niebla"','US (en)','Apila 6 promesas → dilución "cura-todo"','REF'),
('Naturadika','Intestino','"Di adiós a la barriga hinchada" (Magrifit Detox)','ES-España','Estética sin mecanismo; superable con causa-raíz','A'),
('Baïa Food Co.','Intestino','"Sin hinchazón desde los primeros días"','ES-España','Promesa de velocidad sin mecanismo','A'),
('Mariana Esteban','Intestino','Conspiración anti-médico (advertorial ~500K reach)','ES-España','Es intestino, NO hígado → molde portable a MX','A'),
('Wild Dose','Bloating','Historia personal ("el bloating me volteó la vida")','EN (8.8M)','Depende de 1 founder; no traduce a MX','B'),
('Happy Mammoth','Intestino','Autoridad "científico alemán premiado"','EN/UE','Ángulo genérico/copiable; excluye hombres','B'),
('Galmio','Hígado','"Detox de hígado en 5 días"','EN','Claim quemado, LTV bajo, sin mecanismo','B'),
('Dawtox','Hígado','Ángulo alcohol ("90% del alcohol pasa por tu hígado")','FR','Se encierra en el bebedor; no traslada a MX','C'),
("Nature's Finest",'Hígado','Detox estacional + peso ("abdomen plano = hígado limpio")','DE/IT/ES','Sofisticación baja, commoditizado','C'),
]
r=hr+1
for row in comp:
    for c,v in enumerate(row[:5],1): ws.cell(r,c,v)
    t=row[5]
    if t=='REF':
        cell=ws.cell(r,6,'REF'); cell.alignment=CTR; cell.fill=PatternFill('solid',fgColor='D9D9D9'); cell.border=BORD
    else: tiercell(ws,r,6,t)
    r+=1
borders(ws,hr+1,r-1,5)
widths=[18,12,42,14,40,11]
for c,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(c)].width=w
ws.freeze_panes='A5'

# ---------- TAB 3: AVATARES ----------
ws=wb.create_sheet('Avatares (Tier)')
ws['A1']='AVATARES — Core + Sub-avatares (Tier List)'; ws['A1'].font=TITLE
link(ws,2,1,'15','📄 Fuente: DOC 15 — Retrato del Cliente')
hdr=['Avatar','Tier','Tipo','Deseo eje','Dolor central','Emoción dominante','Notas estratégicas']
hr=4
for c,h in enumerate(hdr,1): ws.cell(hr,c,h)
style_header(ws,hr,len(hdr))
av=[
('Core: Mujer 40+ "hace todo bien"','S','Core','Sentirse ligera y ella misma','Hinchazón + fatiga + hígado graso temido','Alivio + Identidad','Núcleo. Prefiere natural. Compra en MercadoLibre/YouTube/TikTok.'),
('"La Hinchada Crónica"','S','Sub-avatar','Deshincharse / vientre plano','Panza inflamada post-comida ("parezco embarazada")','Alivio + vergüenza','GANCHO #1. El de mayor PI (hinchazón 99.8). Lidera el lanzamiento.'),
('"La del Diagnóstico"','A','Sub-avatar','Bajar sus enzimas / sanar el hígado','Hígado graso + ALT/AST altas','Miedo + esperanza','Enzimas = capa de PRUEBA, no gancho. 2ª oleada / retargeting.'),
('"El Crudo / Comida Pesada"','B','Sub-avatar','Recuperar energía / desintoxicar','Cruda, digestión pesada, cerveza/frituras','Culpa + alivio','HIPÓTESIS a validar. Hombre. Villano azúcar/frituras. No núcleo.'),
]
r=hr+1
for row in av:
    ws.cell(r,1,row[0]); tiercell(ws,r,2,row[1])
    for c,v in enumerate(row[2:],3): ws.cell(r,c,v)
    r+=1
borders(ws,hr+1,r-1,7)
widths=[26,7,12,24,30,20,42]
for c,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(c)].width=w
ws.freeze_panes='A5'

# ---------- TAB 4: DESEOS PI ----------
ws=wb.create_sheet('Deseos (PI Tier)')
ws['A1']='DESEOS / DOLORES — Tier List por Índice de Prioridad (data real)'; ws['A1'].font=TITLE
link(ws,2,1,'00','📄 Fuente: DOC 00 Master Spine (ranking PI del corpus)')
ws.cell(3,1,'Tier: S≥90 · A 80-89 · B 65-79 · C <65  |  PI = frecuencia + resonancia. Pool ES = lanzamiento MX; EN = inteligencia US.').font=Font(italic=True,color='808080')
hdr=['Deseo / Tema','Tier (ES)','PI ES','PI EN','n (ES)','Lectura estratégica']
hr=5
for c,h in enumerate(hdr,1): ws.cell(hr,c,h)
style_header(ws,hr,len(hdr))
piEN={x['theme']:x['PI'] for x in rk['EN']}
nES={x['theme']:x['freq'] for x in rk['ES']}
NOTE={'bloating':'★ GANCHO #1 en MX. En US solo #5.','liver_enzymes':'#1 en US (medicalizado); en MX = prueba, no gancho.',
'belief_natural':'Innegociable en ambos pools.','alcohol_liver':'Bajo en MX → villano a re-anclar (comida/azúcar).',
'fatty_liver':'Fuerte en ambos; la razón "seria" detrás del síntoma.'}
r=hr+1
for x in rk['ES']:
    th=x['theme']; ws.cell(r,1,LABELS.get(th,th)); tiercell(ws,r,2,tier(x['PI']))
    ws.cell(r,3,x['PI']).alignment=CTR; ws.cell(r,4,piEN.get(th,'-')).alignment=CTR; ws.cell(r,5,x['freq']).alignment=CTR
    ws.cell(r,6,NOTE.get(th,'')); r+=1
borders(ws,hr+1,r-1,6)
widths=[30,10,9,9,9,44]
for c,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(c)].width=w
ws.freeze_panes='A6'

# ---------- TAB 5: MECANISMOS ----------
ws=wb.create_sheet('Mecanismos (UMP-UMS-USP)')
ws['A1']='MECANISMOS ÚNICOS — UMP / UMS / USP (Tier List)'; ws['A1'].font=TITLE
link(ws,2,1,'19','📄 Fuente: DOC 19 — UMP/UMS/USP')
hdr=['Capa','Nombre de venta (ES-MX)','Núcleo en una línea','Sustanciación (gate)','Tier']
hr=4
for c,h in enumerate(hdr,1): ws.cell(hr,c,h)
style_header(ws,hr,len(hdr))
mec=[
('UMP (problema oculto)','"EL HÍGADO TAPADO"','Te hinchas no por lo que comes, sino porque tu hígado dejó de drenar la bilis y todo se estanca','HIPÓTESIS PERSUASIVA (reencuadre creíble, no diagnóstico clínico)','S'),
('UMS (mecanismo solución)','"DESTAPE HEPÁTICO EN 3 LLAVES"','Diente de León+Alcachofa (mueven bilis) · Colina (fabrica bilis) · Glutatión+NAC (limpian la celda hepática)','Ingredientes = HECHO; efecto 1-a-1 sobre hinchazón = HIPÓTESIS','S'),
('USP (resultado que se vende)','"PANZA PLANA DESDE EL HÍGADO, NO DESDE LA DIETA"','Deshínchate y siéntete ligera destapando el hígado, sin renunciar al pan ni vivir a dieta','Deseo = HECHO; atribución al hígado = HIPÓTESIS','S'),
]
r=hr+1
for row in mec:
    for c,v in enumerate(row[:4],1): ws.cell(r,c,v)
    tiercell(ws,r,5,row[4]); ws.row_dimensions[r].height=48; r+=1
borders(ws,hr+1,r-1,5)
ws.cell(r+1,1,'Villano MX: frituras + azúcar + comida pesada (tamales/ponche) — NO alcohol. Compliance: "apoya/ayuda a", nunca "cura".').font=Font(italic=True,color='C00000')
widths=[24,32,46,40,7]
for c,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(c)].width=w
ws.freeze_panes='A5'

# ---------- TAB 6: OPORTUNIDADES ----------
ws=wb.create_sheet('Oportunidades y Mejoras')
ws['A1']='OPORTUNIDADES Y MEJORAS (vacíos de mercado + arbitraje)'; ws['A1'].font=TITLE
link(ws,2,1,'F0','📄 Fuente: FASE 0 — Vacíos de mercado')
hdr=['Oportunidad','Tipo','Por qué es una ventaja','Prioridad']
hr=4
for c,h in enumerate(hdr,1): ws.cell(hr,c,h)
style_header(ws,hr,len(hdr))
opp=[
('Causa-raíz hepática del bloating en español','Ángulo libre','Nadie en español conecta hinchazón → bilis/hígado. Es el solape exacto hígado+intestino del producto.','S'),
('Des-culpabilización ("no es tu disciplina, es tu hígado")','Ángulo libre','Probado en EN, ausente en español. Potente en MX donde la culpa por peso/comida es alta.','S'),
('Villano re-anclado a comida/azúcar mexicana','Mejora','El eje-alcohol (EN/FR) no traslada a MX; el hígado graso metabólico es el villano correcto y está libre.','S'),
('"El bajón de las 2pm" como síntoma-gancho hepático','Ángulo libre','Sin dueño en español; gancho de reconocimiento instantáneo.','A'),
('Mecanismo glutatión/bilis nombrado en metáfora','Mejora','Los competidores dicen "detox"; nadie traduce el mecanismo a lenguaje sentible propietario.','A'),
('Advertorial de mecanismo-conspiración portado a hígado','Formato','Mariana Esteban probó el molde en español para intestino; hígado está virgen.','A'),
('Segmento hombres / adultos jóvenes','Segmento','Casi todos los ángulos apuntan a mujer 45+; el hígado graso masculino está desatendido.','B'),
('Prueba biomarcador ALT/AST en español','Prueba','En EN hay funnels de lab-test; en ES nadie usa "baja tus enzimas" como prueba tangible.','B'),
('Ronda 2: robustecer reseñas de producto MX','Mejora','El pool de reseñas de marketplace MX quedó delgado (196 ML + 371 Amazon); reforzar antes de escalar pauta.','A'),
]
r=hr+1
for row in opp:
    for c,v in enumerate(row[:3],1): ws.cell(r,c,v)
    tiercell(ws,r,4,row[3]); r+=1
borders(ws,hr+1,r-1,4)
widths=[42,14,58,10]
for c,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(c)].width=w
ws.freeze_panes='A5'

out=f"{FOLDER}\\CONSOLIDADO — Research Hígado MX.xlsx"
wb.save(out)
print("Guardado:",out)
print("Tabs:",wb.sheetnames)
